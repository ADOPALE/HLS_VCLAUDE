"""
data_loader.py — Import du fichier Excel OptiFLUX.

Lecture dynamique de tous les onglets, conversion des types horaires,
détection des formules Excel non calculées.
"""

from __future__ import annotations
import datetime
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd

import config

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# CONVERSION HORAIRE
# ---------------------------------------------------------------------------

def time_to_minutes(value: Any) -> Optional[int]:
    """
    Convertit une valeur horaire Excel en minutes depuis minuit.

    Supporte deux formats :
    - datetime.time → heures × 60 + minutes (+ secondes / 60 arrondies)
    - float [0, 1] → valeur × 1440 arrondie à l'entier le plus proche

    Args:
        value: Valeur brute extraite d'une cellule Excel.

    Returns:
        Entier de minutes depuis minuit, ou None si la valeur est vide/non convertissable.

    Example:
        >>> time_to_minutes(datetime.time(6, 30))
        390
        >>> time_to_minutes(0.25)
        360
    """
    if value is None:
        return None
    if isinstance(value, datetime.time):
        return value.hour * 60 + value.minute + round(value.second / 60)
    if isinstance(value, (int, float)):
        v = float(value)
        if 0.0 <= v <= 1.0:
            return round(v * 1440)
        # Déjà en minutes ?
        if v > 1:
            return int(v)
    return None


def time_to_minutes_strict(value: Any, field_name: str = "") -> int:
    """
    Comme time_to_minutes mais lève ValueError si la valeur est absente.

    Args:
        value: Valeur brute.
        field_name: Nom du champ pour le message d'erreur.

    Returns:
        Entier de minutes.

    Raises:
        ValueError: si la valeur est None ou non convertissable.
    """
    result = time_to_minutes(value)
    if result is None:
        raise ValueError(f"Valeur horaire manquante ou invalide pour le champ '{field_name}' : {value!r}")
    return result


# ---------------------------------------------------------------------------
# CHARGEMENT DU FICHIER
# ---------------------------------------------------------------------------

def load_workbook_data(filepath: str | Path) -> Tuple[openpyxl.Workbook, Dict[str, Any]]:
    """
    Charge le classeur Excel et retourne le workbook + un dict de métadonnées d'import.

    Args:
        filepath: Chemin vers le fichier .xlsx.

    Returns:
        Tuple (workbook openpyxl, dict d'informations d'import avec clés :
            'errors' : list de dicts {type, sheet, row, col, message}
            'warnings' : list de dicts idem
            'formulas' : list de dicts {sheet, row, col, cell_ref, value}
        )

    Raises:
        FileNotFoundError: si le fichier n'existe pas.
        ValueError: si le fichier n'est pas un xlsx valide.
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Fichier introuvable : {filepath}")

    try:
        wb = openpyxl.load_workbook(str(filepath), read_only=False, data_only=True)
    except Exception as exc:
        raise ValueError(f"Impossible de lire le fichier Excel : {exc}") from exc

    meta = {"errors": [], "warnings": [], "formulas": []}
    return wb, meta


def check_required_sheets(wb: openpyxl.Workbook) -> List[str]:
    """
    Vérifie que tous les onglets obligatoires sont présents.

    Args:
        wb: Workbook openpyxl chargé.

    Returns:
        Liste des onglets manquants (vide si OK).
    """
    missing = [s for s in config.REQUIRED_SHEETS if s not in wb.sheetnames]
    return missing


def read_sheet_as_dataframe(wb: openpyxl.Workbook, sheet_name: str) -> pd.DataFrame:
    """
    Lit un onglet Excel en DataFrame pandas.
    La première ligne est utilisée comme en-tête.
    Les lignes entièrement vides sont supprimées.

    Args:
        wb: Workbook openpyxl.
        sheet_name: Nom de l'onglet.

    Returns:
        DataFrame avec les données de l'onglet.
    """
    ws = wb[sheet_name]
    data = list(ws.iter_rows(values_only=True))
    if not data:
        return pd.DataFrame()
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(data[0])]
    rows = data[1:]
    df = pd.DataFrame(rows, columns=headers)
    df = df.dropna(how="all")
    return df


def check_required_columns(df: pd.DataFrame, sheet_name: str) -> List[str]:
    """
    Vérifie que les colonnes obligatoires sont présentes dans le DataFrame.

    Args:
        df: DataFrame de l'onglet.
        sheet_name: Nom de l'onglet (pour lookup dans config.REQUIRED_COLUMNS).

    Returns:
        Liste des colonnes manquantes.
    """
    required = config.REQUIRED_COLUMNS.get(sheet_name, [])
    return [c for c in required if c not in df.columns]


# ---------------------------------------------------------------------------
# LECTURE PARAM RH
# ---------------------------------------------------------------------------

def parse_param_rh(wb: openpyxl.Workbook) -> Dict[str, int]:
    """
    Lit l'onglet 'param RH' et retourne les paramètres en minutes.

    Args:
        wb: Workbook chargé.

    Returns:
        Dict avec clés 'vacation_duration', 'pause_duration', 'start_min', 'end_max'
        (toutes en minutes entières depuis minuit).

    Raises:
        ValueError: si un paramètre est manquant ou non convertissable.
    """
    df = read_sheet_as_dataframe(wb, "param RH")
    if df.empty or len(df) < 1:
        raise ValueError("L'onglet 'param RH' est vide.")
    row = df.iloc[0]
    return {
        "vacation_duration": time_to_minutes_strict(row.get("Format horaire"), "Format horaire"),
        "pause_duration":    time_to_minutes_strict(row.get("Pause"), "Pause"),
        "start_min":         time_to_minutes_strict(row.get("heure début mini"), "heure début mini"),
        "end_max":           time_to_minutes_strict(row.get("heure fin max"), "heure fin max"),
    }


# ---------------------------------------------------------------------------
# LECTURE PARAM SITES
# ---------------------------------------------------------------------------

def parse_param_sites(wb: openpyxl.Workbook) -> Dict[str, Dict[str, Any]]:
    """
    Lit l'onglet 'param Sites' et retourne un dict de sites.

    Args:
        wb: Workbook chargé.

    Returns:
        Dict {libelle_site: {adresse, presence_quai, compat_vehicules: {type: bool}}}
    """
    df = read_sheet_as_dataframe(wb, "param Sites")
    vehicule_cols = [c for c in df.columns if c not in ("Libellé", "Adresses", "Présence de quai")]
    sites = {}
    for _, row in df.iterrows():
        libelle = str(row["Libellé"]).strip() if row["Libellé"] else None
        if not libelle:
            continue
        compat = {}
        for vc in vehicule_cols:
            val = row.get(vc)
            if val is not None:
                compat[str(vc).strip()] = str(val).strip().upper() == "OUI"
        sites[libelle] = {
            "adresse": str(row.get("Adresses", "")).strip(),
            "presence_quai": str(row.get("Présence de quai", "NON")).strip().upper() == "OUI",
            "compat_vehicules": compat,
        }
    return sites


# ---------------------------------------------------------------------------
# LECTURE PARAM VÉHICULES
# ---------------------------------------------------------------------------

def parse_param_vehicules(wb: openpyxl.Workbook) -> Dict[str, Dict[str, Any]]:
    """
    Lit l'onglet 'param Véhicules' et retourne un dict de véhicules.

    Args:
        wb: Workbook chargé.

    Returns:
        Dict {type_vehicule: {stationnement, longueur, largeur, hauteur, poids_max,
                              consommation, cout_carburant, cout_carbone, hayon,
                              compat_contenants, temps_mise_quai, manu_sans_quai,
                              manu_avec_quai}}
    """
    df = read_sheet_as_dataframe(wb, "param Véhicules")
    # Colonnes fixes
    fixed_cols = {
        "Types", "Stationnement initial",
        "dim longueur interne (m)", "dim largeur interne (m)", "dim hauteur interne (m)",
        "Poids max chargement", "Consommation (L/km)", "Cout carburant (€/km)",
        "Cout carbone (kg/km)", "Présence hayon",
        "Temps de mise à quai - manœuvre, contact/admin (minutes)",
        "Manutention sans quai (minutes / contenants)",
        "Manutention avec quai (minutes / contenants)",
    }
    contenant_cols = [c for c in df.columns if c not in fixed_cols]

    vehicules = {}
    for _, row in df.iterrows():
        vtype = str(row.get("Types", "")).strip()
        if not vtype:
            continue

        # Manutention sans quai : peut être "NC"
        manu_sq_raw = row.get("Manutention sans quai (minutes / contenants)")
        if str(manu_sq_raw).strip().upper() == "NC":
            manu_sans_quai = None  # incompatible sites sans quai
        else:
            manu_sans_quai = time_to_minutes(manu_sq_raw)
            # Convertir secondes en minutes si c'était datetime.time(0,0,X)
            if manu_sans_quai is not None and manu_sans_quai == 0:
                # datetime.time(0,0,25) → 25 secondes → ~0.42 min → arrondi à 1
                if isinstance(manu_sq_raw, datetime.time) and manu_sq_raw.second > 0:
                    manu_sans_quai = max(1, round(manu_sq_raw.second / 60))

        manu_aq_raw = row.get("Manutention avec quai (minutes / contenants)")
        manu_avec_quai = time_to_minutes(manu_aq_raw)
        if manu_avec_quai is None:
            manu_avec_quai = 0
        elif manu_avec_quai == 0 and isinstance(manu_aq_raw, datetime.time) and manu_aq_raw.second > 0:
            manu_avec_quai = max(1, round(manu_aq_raw.second / 60))

        # Manutention sans quai : même correction secondes
        if manu_sans_quai == 0 and isinstance(manu_sq_raw, datetime.time) and manu_sq_raw.second > 0:
            manu_sans_quai = max(1, round(manu_sq_raw.second / 60))

        temps_quai = time_to_minutes(row.get("Temps de mise à quai - manœuvre, contact/admin (minutes)"))
        if temps_quai is None:
            temps_quai = 10

        compat_contenants = {}
        for cc in contenant_cols:
            val = row.get(cc)
            if val is not None:
                compat_contenants[str(cc).strip()] = str(val).strip().upper() == "OUI"

        vehicules[vtype] = {
            "stationnement_initial": str(row.get("Stationnement initial", "")).strip(),
            "longueur": float(row.get("dim longueur interne (m)", 0) or 0),
            "largeur":  float(row.get("dim largeur interne (m)", 0) or 0),
            "hauteur":  float(row.get("dim hauteur interne (m)", 0) or 0),
            "poids_max": float(row.get("Poids max chargement", 0) or 0),
            "consommation": float(row.get("Consommation (L/km)", 0) or 0),
            "cout_carburant": float(row.get("Cout carburant (€/km)", 0) or 0),
            "cout_carbone": float(row.get("Cout carbone (kg/km)", 0) or 0),
            "hayon": str(row.get("Présence hayon", "NON")).strip().upper() == "OUI",
            "compat_contenants": compat_contenants,
            "temps_mise_quai": temps_quai,
            "manu_sans_quai": manu_sans_quai,
            "manu_avec_quai": manu_avec_quai,
        }
    return vehicules


# ---------------------------------------------------------------------------
# LECTURE PARAM CONTENANTS
# ---------------------------------------------------------------------------

def parse_param_contenants(wb: openpyxl.Workbook) -> Dict[str, Dict[str, Any]]:
    """
    Lit l'onglet 'param Contenants' et retourne un dict de contenants.

    Args:
        wb: Workbook chargé.

    Returns:
        Dict {libelle: {longueur, largeur, poids_vide, poids_plein}}
    """
    df = read_sheet_as_dataframe(wb, "param Contenants")
    contenants = {}
    for _, row in df.iterrows():
        lib = str(row.get("libellé", "")).strip()
        if not lib or lib.lower() in ("none", "nan"):
            continue
        contenants[lib] = {
            "longueur": float(row.get("dim longueur (m)", 0) or 0),
            "largeur":  float(row.get("dim largeur (m)", 0) or 0),
            "poids_vide": float(row.get("Poids vide (T)", 0) or 0),
            "poids_plein": float(row.get("Poids plein (T)", 0) or 0),
        }
    return contenants


# ---------------------------------------------------------------------------
# LECTURE MATRICES
# ---------------------------------------------------------------------------

def parse_matrix(wb: openpyxl.Workbook, sheet_name: str) -> Dict[str, Dict[str, float]]:
    """
    Lit une matrice carrée (durée ou distance) depuis un onglet Excel.
    La première colonne et la première ligne sont les labels de sites.

    Args:
        wb: Workbook chargé.
        sheet_name: 'matrice Durée' ou 'matrice Dist'.

    Returns:
        Dict {site_départ: {site_arrivée: valeur_float}}
    """
    ws = wb[sheet_name]
    data = list(ws.iter_rows(values_only=True))
    if not data:
        return {}
    header = data[0]
    col_sites = [str(h).strip() if h is not None else "" for h in header[1:]]
    matrix = {}
    for row in data[1:]:
        if not row or row[0] is None:
            continue
        row_site = str(row[0]).strip()
        if not row_site:
            continue
        matrix[row_site] = {}
        for j, col_site in enumerate(col_sites):
            if not col_site:
                continue
            val = row[j + 1]
            try:
                matrix[row_site][col_site] = float(val) if val is not None else 0.0
            except (TypeError, ValueError):
                matrix[row_site][col_site] = 0.0
    return matrix


# ---------------------------------------------------------------------------
# LECTURE M FLUX
# ---------------------------------------------------------------------------

def detect_formula_cells(wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    """
    Détecte les cellules de quantité contenant des formules Excel non calculées
    dans l'onglet 'M flux'.

    Une formule non calculée est une chaîne commençant par '=' dans les colonnes
    'Quantité Lundi' à 'Quantité Dimanche'.

    Args:
        wb: Workbook chargé.

    Returns:
        Liste de dicts {sheet, row_excel, col_excel, cell_ref, value}
    """
    if "M flux" not in wb.sheetnames:
        return []
    ws = wb["M flux"]
    data = list(ws.iter_rows(values_only=True))
    if not data:
        return []

    header = data[0]
    qty_cols = list(config.DAY_COLUMNS.values())
    qty_col_indices = {
        i: str(h).strip()
        for i, h in enumerate(header)
        if str(h).strip() in qty_cols
    }

    formulas = []
    for row_idx, row in enumerate(data[1:], start=2):  # row_idx est 1-based Excel
        for col_idx, col_name in qty_col_indices.items():
            val = row[col_idx] if col_idx < len(row) else None
            if isinstance(val, str) and val.strip().startswith("="):
                col_letter = _col_index_to_letter(col_idx + 1)
                formulas.append({
                    "sheet": "M flux",
                    "row_excel": row_idx + 1,
                    "col_excel": col_letter,
                    "cell_ref": f"{col_letter}{row_idx + 1}",
                    "value": val,
                })
    return formulas


def _col_index_to_letter(n: int) -> str:
    """Convertit un index de colonne (1-based) en lettre Excel."""
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result


def parse_m_flux(wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
    """
    Lit l'onglet 'M flux' et retourne la liste brute de tous les flux.

    N'applique pas de filtre par jour ni par nature. Le filtrage est délégué
    à preprocessing.py.

    Args:
        wb: Workbook chargé.

    Returns:
        Liste de dicts avec tous les champs du flux (horaires convertis en minutes).
    """
    df = read_sheet_as_dataframe(wb, "M flux")
    flux_list = []
    col_nature = "Nature du flux (les tournées sont elles à prévoir avec une obligation de transport ou une obligation de passage?)"

    for idx, row in df.iterrows():
        # Horaires
        heure_dispo = time_to_minutes(row.get("Heure de mise à disposition min départ"))
        heure_max = time_to_minutes(row.get("Heure max de livraison à la destination"))

        # Quantités par jour
        quantites = {}
        for day_idx, day_col in config.DAY_COLUMNS.items():
            val = row.get(day_col)
            if val is None or (isinstance(val, float) and pd.isna(val)):
                quantites[day_idx] = 0
            else:
                try:
                    quantites[day_idx] = int(float(val))
                except (TypeError, ValueError):
                    quantites[day_idx] = 0

        transport_mixte_raw = str(row.get("Transport mixte possible (OUI / NON)", "NON")).strip().upper()
        tournee_raw = str(row.get("Tournées mutualisées ? (OUI / NON)", "NON")).strip().upper()
        urgence_raw = str(row.get("Urgence / flux prioritaire \n(Oui/Non)", "Non")).strip().upper()

        flux_list.append({
            "id_flux": int(idx) + 2,  # +2 : ligne Excel (header=1, data commence à 2)
            "site_depart": str(row.get("Point de départ", "")).strip(),
            "site_arrivee": str(row.get("Point de destination", "")).strip(),
            "fonction_support": str(row.get("Fonction Support associée", "")).strip(),
            "nature_flux": str(row.get("Nature du Flux \n(champ libre)", "") or "").strip(),
            "type_contenant": str(row.get("Nature de contenant", "")).strip(),
            "statut_plein_vide": str(row.get("Plein / vide", "")).strip(),
            "statut_propre_sale": str(row.get("Sale / propre", "")).strip(),
            "aller_retour": str(row.get("Aller/Retour", "") or "").strip(),
            "transport_mixte": transport_mixte_raw == "OUI",
            "regle_exclusion": row.get("Règles d'exclusions si transport mixte") or None,
            "tournee_mutualisee": tournee_raw == "OUI",
            "nom_tournee": row.get("Nom de la tournée mutualisée le cas échéant") or None,
            "nature": str(row.get(col_nature, "") or "").strip(),
            "quantites": quantites,
            "heure_dispo": heure_dispo,
            "heure_max_livraison": heure_max,
            "urgent": urgence_raw in ("OUI", "OUI "),
        })
    return flux_list


# ---------------------------------------------------------------------------
# POINT D'ENTRÉE PRINCIPAL
# ---------------------------------------------------------------------------

def load_all(filepath: str | Path) -> Dict[str, Any]:
    """
    Charge et parse l'intégralité du fichier Excel.

    Args:
        filepath: Chemin vers le fichier .xlsx.

    Returns:
        Dict avec clés :
            'rh'         : paramètres RH (dict)
            'sites'      : sites (dict)
            'vehicules'  : véhicules (dict)
            'contenants' : contenants (dict)
            'matrix_dur' : matrice durées (dict)
            'matrix_dist': matrice distances (dict)
            'flux_brut'  : liste brute des flux (list)
            'errors'     : erreurs bloquantes (list)
            'warnings'   : alertes non bloquantes (list)
            'formulas'   : formules non calculées (list)

    Raises:
        FileNotFoundError: fichier introuvable.
        ValueError: format invalide.
    """
    wb, meta = load_workbook_data(filepath)
    errors = meta["errors"]
    warnings = meta["warnings"]

    # 1. Onglets obligatoires
    missing_sheets = check_required_sheets(wb)
    for s in missing_sheets:
        errors.append({
            "type": "ONGLET_MANQUANT",
            "sheet": s,
            "row": None,
            "col": None,
            "message": f"Onglet obligatoire manquant : '{s}'",
        })

    if missing_sheets:
        return {"errors": errors, "warnings": warnings, "formulas": []}

    # 2. Colonnes obligatoires
    for sheet_name in config.REQUIRED_COLUMNS:
        if sheet_name not in wb.sheetnames:
            continue
        df = read_sheet_as_dataframe(wb, sheet_name)
        missing_cols = check_required_columns(df, sheet_name)
        for c in missing_cols:
            errors.append({
                "type": "COLONNE_MANQUANTE",
                "sheet": sheet_name,
                "row": 1,
                "col": c,
                "message": f"Colonne obligatoire manquante dans '{sheet_name}' : '{c}'",
            })

    # 3. Formules Excel non calculées
    formulas = detect_formula_cells(wb)
    if formulas:
        errors.append({
            "type": "FORMULES_NON_CALCULEES",
            "sheet": "M flux",
            "row": None,
            "col": None,
            "message": f"{len(formulas)} cellule(s) de quantité contiennent des formules non calculées.",
            "detail": formulas,
        })

    if errors:
        return {"errors": errors, "warnings": warnings, "formulas": formulas}

    # 4. Parsing des onglets
    try:
        rh = parse_param_rh(wb)
    except ValueError as e:
        errors.append({"type": "PARSE_ERROR", "sheet": "param RH", "message": str(e)})
        rh = config.DEFAULT_RH.copy()

    sites = parse_param_sites(wb)
    vehicules = parse_param_vehicules(wb)
    contenants = parse_param_contenants(wb)
    matrix_dur = parse_matrix(wb, "matrice Durée")
    matrix_dist = parse_matrix(wb, "matrice Dist")
    flux_brut = parse_m_flux(wb)

    return {
        "rh": rh,
        "sites": sites,
        "vehicules": vehicules,
        "contenants": contenants,
        "matrix_dur": matrix_dur,
        "matrix_dist": matrix_dist,
        "flux_brut": flux_brut,
        "errors": errors,
        "warnings": warnings,
        "formulas": formulas,
    }
